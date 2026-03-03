import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import pdfplumber
import os
import re
import unicodedata
from datetime import datetime

# --- BIBLIOTECAS DO EXCEL E GOOGLE SHEETS ---
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import gspread
from oauth2client.service_account import ServiceAccountCredentials

# --- VARIÁVEIS GLOBAIS DE CAMINHO E ID DA NUVEM ---
caminho_estoque_global = ""
caminho_pedidos_global = ""
caminho_datas_global = ""
ID_PLANILHA_GOOGLE = "191lthbx21pdEpPF_SZJAMgOviXE-e8HnrWHdUzxzjMc"

def limpar_nome(nome):
    if not nome or pd.isna(nome): return ""
    return re.sub(r'\s+', ' ', str(nome)).strip().upper()

def limpar_para_busca_agressiva(texto):
    if not isinstance(texto, str): return ""
    texto = str(texto).upper()
    texto = unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode('utf-8')
    texto = re.sub(r'\b(EMP|EMPENHO|PREFEITURA|MUNICIPAL|CAMARA|FUNDO|DE|DA|DO|LTDA|ME)\b', ' ', texto)
    texto = re.sub(r'\d+', '', texto)
    return re.sub(r'[^A-Z]', '', texto)

def ler_planilha_datas(caminho_excel):
    if not caminho_excel or "Nenhum" in caminho_excel: return {}
    try:
        def extrair_dicionario(df_cru):
            header_idx = -1
            for i, row in df_cru.iterrows():
                row_str = " ".join([str(x).upper() for x in row.values])
                if "CIDADE" in row_str and "LINK" in row_str:
                    header_idx = i
                    break
            if header_idx == -1: return {}
            
            headers = [str(x).upper() for x in df_cru.iloc[header_idx].values]
            idx_cidade = next((i for i, h in enumerate(headers) if "CIDADE" in h), -1)
            idx_link = next((i for i, h in enumerate(headers) if "LINK" in h), -1)
            idx_data = next((i for i, h in enumerate(headers) if "EMP" in h), -1)
            if idx_cidade == -1: return {}
            
            infos = {}
            for i in range(header_idx + 1, len(df_cru)):
                row = df_cru.iloc[i]
                nome_cidade = limpar_nome(row.iloc[idx_cidade]) if pd.notnull(row.iloc[idx_cidade]) else ""
                if not nome_cidade or nome_cidade == "CIDADE": continue
                
                data_val = row.iloc[idx_data] if idx_data != -1 else None
                link_val = row.iloc[idx_link] if idx_link != -1 else ""
                
                if pd.notnull(data_val) and str(data_val).strip() != "" and str(data_val).upper() != "EMPENHO":
                    if isinstance(data_val, str):
                        try: data_obj = pd.to_datetime(data_val, dayfirst=True)
                        except: data_obj = datetime.max
                    else: data_obj = pd.to_datetime(data_val)
                else: data_obj = datetime.max
                    
                infos[nome_cidade] = {"data": data_obj, "link": str(link_val).strip() if pd.notnull(link_val) else ""}
            return infos

        if caminho_excel.lower().endswith('.csv'): return extrair_dicionario(pd.read_csv(caminho_excel, header=None))
        else:
            xls = pd.ExcelFile(caminho_excel)
            abas = xls.sheet_names
            aba_alvo = next((aba for aba in abas if "EMPENHO" in aba.upper() and "CÓPIA" not in aba.upper()), None)
            if not aba_alvo: aba_alvo = next((aba for aba in abas if "EMPENHO" in aba.upper()), None)
            if aba_alvo:
                dict_res = extrair_dicionario(pd.read_excel(xls, sheet_name=aba_alvo, header=None))
                if dict_res: return dict_res
            for aba in abas:
                dict_res = extrair_dicionario(pd.read_excel(xls, sheet_name=aba, header=None))
                if dict_res: return dict_res
            return {}
    except Exception as e:
        return {}

def extrair_estoque_pdf(caminho_pdf):
    estoque = {} 
    with pdfplumber.open(caminho_pdf) as pdf:
        for pagina in pdf.pages:
            texto = pagina.extract_text()
            if not texto: continue
            for linha in texto.split('\n'):
                partes = linha.strip().split()
                if len(partes) >= 5 and partes[0].isdigit():
                    codigo_erp = partes[0]
                    num_count = 0
                    for p in reversed(partes):
                        if re.match(r'^-?[\d.,]+$', p): num_count += 1
                        else: break
                    
                    if num_count >= 1:
                        str_qtd = partes[-num_count] 
                        str_preco = "0,00"
                        if num_count >= 3: 
                            str_qtd = partes[-3] if num_count == 3 else partes[-4]
                            str_preco = partes[-2]
                        elif num_count == 2:
                            str_qtd = partes[-2]
                            str_preco = partes[-1]
                            
                        nome_estoque = ""
                        fim_nome = len(partes) - num_count
                        if fim_nome > 0 and partes[fim_nome - 1] in ["UN", "PC", "CX", "KG", "MT", "PÇ"]:
                            fim_nome -= 1
                        if fim_nome > 1:
                            nome_estoque = " ".join(partes[1:fim_nome]).strip()
                            
                        try:
                            val_str = str_qtd
                            if ',' in val_str and '.' in val_str: val_str = val_str.replace('.', '').replace(',', '.')
                            elif ',' in val_str: val_str = val_str.replace(',', '.')
                            qtd_val = int(float(val_str))
                            
                            p_str = str_preco
                            if ',' in p_str and '.' in p_str: p_str = p_str.replace('.', '').replace(',', '.')
                            elif ',' in p_str: p_str = p_str.replace(',', '.')
                            preco_val = float(p_str)
                            
                            estoque[codigo_erp] = {'qtd': qtd_val, 'preco': preco_val, 'nome': nome_estoque}
                        except ValueError:
                            pass
    return estoque

def extrair_pedidos_pdf(caminho_pdf, estoque_conhecido, compras_transito):
    codigos_validos = set(list(estoque_conhecido.keys()) + list(compras_transito.keys()))
    empenhos = {}
    current_empenho = "Pedidos Avulsos"
    
    with pdfplumber.open(caminho_pdf) as pdf:
        for num_pag, pagina in enumerate(pdf.pages):
            texto = pagina.extract_text()
            if not texto: continue
            
            for linha in texto.split('\n'):
                linha_upper = limpar_nome(linha)
                if not linha_upper: continue
                
                if any(x in linha_upper for x in ["SUBTOTAL", "FRETE", "TOTAL GERAL", "DESCONTO"]):
                    continue
                    
                partes = linha_upper.split()
                if len(partes) < 3: continue
                
                # --- CABEÇALHO DO COMPUTADOR ---
                if ("COMPUTADOR" in linha_upper or "TPLAN" in linha_upper) and partes[0].isdigit():
                    idx_comp = next((i for i, p in enumerate(partes) if "COMPUTADOR" in p or "TPLAN" in p), -1)
                    
                    if idx_comp > 2:
                        cidade_ref = " ".join(partes[2:idx_comp])
                    else:
                        cidade_ref = " ".join(partes[1:idx_comp]) if idx_comp > 1 else f"Pag_{num_pag+1}"
                    
                    match_id = re.search(r'(TPLAN\s*#?[^\s]+)', linha_upper)
                    pc_id = match_id.group(1).strip() if match_id else ""
                    
                    current_empenho = f"[{cidade_ref}] - {pc_id}" if pc_id else f"[{cidade_ref}]"
                    
                    try: valor_float = float(partes[-1].replace('.', '').replace(',', '.'))
                    except: valor_float = 0.0
                    
                    qtd_pcs = 1
                    for i in range(len(partes)-1, max(0, len(partes)-4), -1):
                        if re.match(r'^\d+$', partes[i]):
                            qtd_pcs = int(partes[i])
                            break
                    
                    # Cria o grupo do Empenho, mas NÃO ADICIONA o Computador como peça!
                    if current_empenho not in empenhos:
                        empenhos[current_empenho] = {"cidade_ref": cidade_ref, "valor_total": valor_float, "qtd_pcs": qtd_pcs, "pecas": {}}
                    else:
                        empenhos[current_empenho]["valor_total"] += valor_float
                        empenhos[current_empenho]["qtd_pcs"] += qtd_pcs
                    continue # Salta para a próxima linha sem adicionar nada à lista de compras
                
                # --- INSUMOS E PEÇAS REAIS ---
                if partes[0].isdigit() and (partes[1].isdigit() or len(partes[1]) >= 4):
                    codigo_encontrado = None
                    for p in partes[:3]:
                        if p in codigos_validos:
                            codigo_encontrado = p
                            break
                    if not codigo_encontrado:
                        codigo_encontrado = partes[1] if partes[1].isdigit() else partes[0]
                        
                    qtd = 0
                    fim_nome = len(partes) - 1
                    
                    for i in range(len(partes)-1, max(0, len(partes)-6), -1):
                        p = partes[i]
                        if re.match(r'^-?[\d\.]+,\d{2}$', p): 
                            fim_nome = min(fim_nome, i - 1)
                            continue
                        if p in ["UN", "PC", "CX", "PÇ", "MT", "KG"]: 
                            fim_nome = min(fim_nome, i - 1)
                            continue
                        if re.match(r'^\d+$', p) and qtd == 0: 
                            qtd = int(p)
                            fim_nome = min(fim_nome, i - 1)
                            continue
                        break 
                        
                    if qtd == 0: qtd = 1
                        
                    try: inicio = partes.index(codigo_encontrado) + 1
                    except: inicio = 2
                        
                    if inicio <= fim_nome:
                        words = partes[inicio:fim_nome+1]
                        meio_idx = len(words) // 2
                        str1 = re.sub(r'\W+', '', " ".join(words[:meio_idx]))
                        str2 = re.sub(r'\W+', '', " ".join(words[meio_idx:]))
                        if str1 == str2 and len(str1) > 2: 
                            nome_final = " ".join(words[:meio_idx])
                        else: 
                            nome_final = " ".join(words)
                    else:
                        nome_final = "ITEM DESCONHECIDO"
                        
                    if current_empenho not in empenhos:
                        empenhos[current_empenho] = {"cidade_ref": "Desconhecido", "valor_total": 0.0, "qtd_pcs": 1, "pecas": {}}
                        
                    if codigo_encontrado in empenhos[current_empenho]["pecas"]:
                        empenhos[current_empenho]["pecas"][codigo_encontrado]["qtd"] += qtd
                    else:
                        empenhos[current_empenho]["pecas"][codigo_encontrado] = {"qtd": qtd, "nome": nome_final}
    return empenhos

def obter_planilha_nuvem():
    escopo = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    try:
        credenciais = ServiceAccountCredentials.from_json_keyfile_name("credenciais.json", escopo)
    except:
        raise Exception("ERRO: Arquivo 'credenciais.json' não encontrado na pasta do programa.")
    cliente = gspread.authorize(credenciais)
    return cliente.open_by_key(ID_PLANILHA_GOOGLE)

def ler_estoque_anterior(planilha):
    try:
        aba = planilha.worksheet("Estoque Inicial")
        dados = aba.get_all_values()
        if len(dados) <= 1: return {}
        headers = [str(h).strip().upper() for h in dados[0]]
        idx_item = next((i for i, h in enumerate(headers) if "ITEM" in h), -1)
        idx_fisico = next((i for i, h in enumerate(headers) if "FÍSICO" in h or "FISICO" in h), -1)
        
        estoque_ant = {}
        if idx_item != -1 and idx_fisico != -1:
            for row in dados[1:]:
                if len(row) > max(idx_item, idx_fisico):
                    item_str = str(row[idx_item]).strip()
                    cod = item_str.split(" - ")[0].strip()
                    try:
                        qtd = int(row[idx_fisico])
                        estoque_ant[cod] = qtd
                    except: pass
        return estoque_ant
    except gspread.exceptions.WorksheetNotFound:
        return {}

def processar_auto_baixa_transito(planilha, estoque_atual, estoque_anterior, salvar_nuvem=True):
    try: aba = planilha.worksheet("Compras em Trânsito")
    except: 
        aba = planilha.add_worksheet(title="Compras em Trânsito", rows="100", cols="5")
        aba.update([["Código", "Qtd Comprada", "Descrição do Item (Opcional)"]], value_input_option="USER_ENTERED")
        return {}
        
    dados = aba.get_all_values()
    if len(dados) <= 1: return {}
    
    headers = [str(h).strip().upper() for h in dados[0]]
    idx_cod = next((i for i, h in enumerate(headers) if "CÓD" in h or "COD" in h), -1)
    idx_qtd = next((i for i, h in enumerate(headers) if "QTD" in h or "QUANT" in h or "COMPRADA" in h), -1)
    
    compras_ativas = {}
    novas_linhas = [dados[0]] 
    houve_alteracao = False
    
    if idx_cod != -1 and idx_qtd != -1:
        for row in dados[1:]:
            while len(row) < 3: row.append("") 
            
            cod = str(row[idx_cod]).strip()
            try: qtd_comprada = int(str(row[idx_qtd]).strip())
            except: qtd_comprada = 0
            
            if cod and qtd_comprada > 0:
                fisico_hoje = estoque_atual.get(cod, {}).get('qtd', 0)
                fisico_ontem = estoque_anterior.get(cod, 0)
                
                if fisico_hoje > fisico_ontem:
                    chegou = fisico_hoje - fisico_ontem
                    baixa = min(chegou, qtd_comprada) 
                    qtd_comprada -= baixa
                    estoque_anterior[cod] += baixa 
                    houve_alteracao = True
                    
            if cod and qtd_comprada > 0:
                row[idx_qtd] = qtd_comprada
                novas_linhas.append(row)
                compras_ativas[cod] = compras_ativas.get(cod, 0) + qtd_comprada
                
    if houve_alteracao and salvar_nuvem:
        aba.clear()
        aba.update(novas_linhas, value_input_option="USER_ENTERED")
        
    return compras_ativas

def processar_cruzamento_dados(estoque_atual, demanda_empenhos, infos_conhecidas, modo_prioridade, compras_transito, salvar_na_nuvem):
    mapa_descricoes = {}
    
    for cod, dados in estoque_atual.items():
        if 'nome' in dados and dados['nome']:
            mapa_descricoes[cod] = f"{cod} - {dados['nome']}"
            
    for emp_id, data in demanda_empenhos.items():
        for cod, dados in data["pecas"].items():
            if cod not in mapa_descricoes:
                mapa_descricoes[cod] = f"{cod} - {dados['nome']}"
            if cod not in estoque_atual:
                estoque_atual[cod] = {'qtd': 0, 'preco': 0.0, 'nome': dados['nome']}
                
    for cod in compras_transito.keys():
        if cod not in estoque_atual:
            estoque_atual[cod] = {'qtd': 0, 'preco': 0.0, 'nome': 'CADASTRADO NO TRÂNSITO'}
        if cod not in mapa_descricoes:
            mapa_descricoes[cod] = f"{cod} - {estoque_atual[cod].get('nome', '')}"

    estoque_linhas = []
    for cod, dados in estoque_atual.items():
        fisico = dados['qtd']
        transito = compras_transito.get(cod, 0)
        total_disponivel = fisico + transito
        
        estoque_linhas.append({
            "Item": mapa_descricoes.get(cod, str(cod)),
            "Quantidade Inicial": total_disponivel,
            "Preço Unitário": dados['preco'],
            "Físico (PDF)": fisico,
            "Em Trânsito (Sheets)": transito
        })
        
    estoque_linhas = sorted(estoque_linhas, key=lambda x: str(x["Item"]))
    df_estoque = pd.DataFrame(estoque_linhas)
    
    fila_empenhos = []
    for emp_id, data in demanda_empenhos.items():
        cidade_pdf_limpa = limpar_para_busca_agressiva(data["cidade_ref"])
        info_cidade = {"data": datetime.max, "link": "Não encontrado"}
        for cidade_excel, info_ex in infos_conhecidas.items():
            cidade_excel_limpa = limpar_para_busca_agressiva(cidade_excel)
            if len(cidade_pdf_limpa) > 2 and len(cidade_excel_limpa) > 2:
                if cidade_pdf_limpa in cidade_excel_limpa or cidade_excel_limpa in cidade_pdf_limpa:
                    info_cidade = info_ex
                    break
        fila_empenhos.append({"id": emp_id, "valor_total": data["valor_total"], "qtd_pcs": data["qtd_pcs"], "data_empenho": info_cidade["data"], "link": info_cidade["link"], "pecas": data["pecas"]})
        
    estoque_virtual = {cod: (dados['qtd'] + compras_transito.get(cod, 0)) for cod, dados in estoque_atual.items()}
    empenhos_ordenados = []
    ordem_separacao = 1
    hoje = datetime.now()
    
    while fila_empenhos:
        for emp in fila_empenhos:
            faltantes = 0
            for cod, dados in emp["pecas"].items():
                saldo = estoque_virtual.get(cod, 0)
                if modo_prioridade == "unidades":
                    if saldo <= 0: faltantes += dados["qtd"]
                    else: faltantes += max(0, dados["qtd"] - saldo)
                else:
                    if saldo < dados["qtd"]: faltantes += 1
            emp["score_faltantes"] = faltantes
            idade_dias = (hoje - emp["data_empenho"]).days if emp["data_empenho"] != datetime.max else 0
            if idade_dias > 20:
                grupo = 0 
                chave3, chave4 = emp["data_empenho"].toordinal(), -emp["valor_total"]
            else:
                grupo = 1 
                chave3, chave4 = -emp["valor_total"], emp["data_empenho"].toordinal()
            emp["sort_key"] = (grupo, emp["score_faltantes"], chave3, chave4)
            
        fila_empenhos.sort(key=lambda x: x["sort_key"])
        emp_da_vez = fila_empenhos.pop(0)
        for cod, dados in emp_da_vez["pecas"].items():
            estoque_virtual[cod] = estoque_virtual.get(cod, 0) - dados["qtd"]
        emp_da_vez["ordem_inicial"] = ordem_separacao
        ordem_separacao += 1
        empenhos_ordenados.append(emp_da_vez)

    resultado_separacao = []
    for emp in empenhos_ordenados:
        data_formatada = emp["data_empenho"].strftime("%d/%m/%Y") if emp["data_empenho"] != datetime.max else "Sem data"
        for codigo_peca, dados_peca in emp["pecas"].items():
            resultado_separacao.append({
                "Prioridade (Fila)": emp["ordem_inicial"],
                "Link": emp["link"],
                "Data do Empenho": data_formatada,
                "Valor do Pedido": emp["valor_total"],
                "Qtd. Máquinas": emp["qtd_pcs"],
                "Empenho": emp["id"],
                "Item": mapa_descricoes.get(codigo_peca, f"{codigo_peca} - {dados_peca['nome']}"),
                "Qtd. Solicitada": dados_peca["qtd"],
                "Qtd. a Comprar": "", 
                "Saldo Estimado": ""
            })

    df_separacao = pd.DataFrame(resultado_separacao)
    
    for i in range(len(df_separacao)):
        row = i + 2
        if salvar_na_nuvem:
            formula_saldo = f'=SEERRO(PROCV(G{row}; \'Estoque Inicial\'!A:B; 2; FALSO); 0) - SOMASES(INDIRETO("H2:H"&LIN()); INDIRETO("G2:G"&LIN()); G{row})'
            formula_comprar = f'=SE(J{row}<0; MÍNIMO(H{row}; ABS(J{row})); 0)'
        else:
            formula_saldo = f'=IFERROR(VLOOKUP(G{row}, \'Estoque Inicial\'!A:B, 2, FALSE), 0) - SUMIFS($H$2:$H${row}, $G$2:$G${row}, G{row})'
            formula_comprar = f'=IF(J{row}<0, MIN(H{row}, ABS(J{row})), 0)'
            
        df_separacao.at[i, "Saldo Estimado"] = formula_saldo
        df_separacao.at[i, "Qtd. a Comprar"] = formula_comprar
        
    itens_solicitados = df_separacao["Item"].unique()
    compras_rows = []
    ridx = 2
    
    for item in itens_solicitados:
        codigo_peca = str(item).split(" - ")[0]
        preco_base = estoque_atual.get(codigo_peca, {}).get('preco', 0.0)
        
        if salvar_na_nuvem:
            form_total_solicitado = f"=SOMASES('Separação por Empenho'!H:H; 'Separação por Empenho'!G:G; A{ridx})"
            form_qtd_comprar = f"=SOMASES('Separação por Empenho'!I:I; 'Separação por Empenho'!G:G; A{ridx})"
        else:
            form_total_solicitado = f"=SUMIFS('Separação por Empenho'!H:H, 'Separação por Empenho'!G:G, A{ridx})"
            form_qtd_comprar = f"=SUMIFS('Separação por Empenho'!I:I, 'Separação por Empenho'!G:G, A{ridx})"
            
        form_custo = f"=C{ridx}*D{ridx}"
            
        compras_rows.append({
            "Item": item,
            "Total Solicitado em Pedidos": form_total_solicitado,
            "Qtd. a Comprar": form_qtd_comprar, 
            "Preço Unit. Base": preco_base,
            "Custo Estimado": form_custo
        })
        ridx += 1
            
    if not compras_rows: compras_rows.append({"Item": "ESTOQUE SUFICIENTE", "Total Solicitado em Pedidos": 0, "Qtd. a Comprar": 0, "Preço Unit. Base": 0.0, "Custo Estimado": 0.0})
    df_compras = pd.DataFrame(compras_rows)
    return df_separacao, df_compras, df_estoque

def exportar_google_sheets(planilha, df_separacao, df_compras, df_estoque):
    df_separacao = df_separacao.astype(object).fillna("")
    df_compras = df_compras.astype(object).fillna("")
    df_estoque = df_estoque.astype(object).fillna("")

    def atualizar_aba_e_formatar(nome_aba, df):
        try: worksheet = planilha.worksheet(nome_aba)
        except: worksheet = planilha.add_worksheet(title=nome_aba, rows="100", cols="20")
        
        worksheet.clear()
        dados_lista = [df.columns.values.tolist()] + df.values.tolist()
        worksheet.update(dados_lista, value_input_option="USER_ENTERED")

        try:
            sheet_id = worksheet.id
            cols_count = len(df.columns)
            requests = []
            
            requests.append({
                "repeatCell": {
                    "range": {"sheetId": sheet_id, "startRowIndex": 1, "endRowIndex": len(df) + 1, "startColumnIndex": 0, "endColumnIndex": cols_count},
                    "cell": {"userEnteredFormat": {"backgroundColor": {"red": 1.0, "green": 1.0, "blue": 1.0}}},
                    "fields": "userEnteredFormat.backgroundColor"
                }
            })
            
            requests.append({
                "repeatCell": {
                    "range": {"sheetId": sheet_id, "startRowIndex": 0, "endRowIndex": 1, "startColumnIndex": 0, "endColumnIndex": cols_count},
                    "cell": {
                        "userEnteredFormat": {
                            "backgroundColor": {"red": 0.12, "green": 0.30, "blue": 0.47},
                            "textFormat": {"foregroundColor": {"red": 1.0, "green": 1.0, "blue": 1.0}, "bold": True},
                            "horizontalAlignment": "CENTER"
                        }
                    },
                    "fields": "userEnteredFormat(backgroundColor,textFormat,horizontalAlignment)"
                }
            })
            
            if nome_aba == "Separação por Empenho":
                empenho_atual = None
                usar_cinza = False
                start_gray_idx = None
                
                for idx, row_val in enumerate(df["Empenho"]):
                    if row_val != empenho_atual:
                        usar_cinza = not usar_cinza
                        empenho_atual = row_val
                        
                    if usar_cinza:
                        if start_gray_idx is None: start_gray_idx = idx
                    else:
                        if start_gray_idx is not None:
                            requests.append({
                                "repeatCell": {
                                    "range": {"sheetId": sheet_id, "startRowIndex": start_gray_idx + 1, "endRowIndex": idx + 1, "startColumnIndex": 0, "endColumnIndex": cols_count},
                                    "cell": {"userEnteredFormat": {"backgroundColor": {"red": 0.95, "green": 0.95, "blue": 0.95}}},
                                    "fields": "userEnteredFormat.backgroundColor"
                                }
                            })
                            start_gray_idx = None
                            
                if start_gray_idx is not None:
                    requests.append({
                        "repeatCell": {
                            "range": {"sheetId": sheet_id, "startRowIndex": start_gray_idx + 1, "endRowIndex": len(df) + 1, "startColumnIndex": 0, "endColumnIndex": cols_count},
                            "cell": {"userEnteredFormat": {"backgroundColor": {"red": 0.95, "green": 0.95, "blue": 0.95}}},
                            "fields": "userEnteredFormat.backgroundColor"
                        }
                    })
            
            if requests:
                planilha.batch_update({"requests": requests})
        except Exception as e:
            pass

    atualizar_aba_e_formatar("Separação por Empenho", df_separacao)
    atualizar_aba_e_formatar("Resumo para Compras", df_compras)
    atualizar_aba_e_formatar("Estoque Inicial", df_estoque)

def formatar_excel(writer, nome_aba):
    workbook = writer.book
    worksheet = writer.sheets[nome_aba]
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    borda_fina = Border(left=Side(style='thin', color="D3D3D3"), right=Side(style='thin', color="D3D3D3"), top=Side(style='thin', color="D3D3D3"), bottom=Side(style='thin', color="D3D3D3"))
    
    for cell in worksheet["1:1"]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = borda_fina
        
    for col in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        nome_coluna = str(col[0].value) if col[0].value else ""
        
        for cell in col:
            if cell.row > 1 and cell.value is not None:
                if nome_coluna in ["Preço Unit. Base", "Custo Estimado", "Preço Unitário", "Valor do Pedido"]:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '"R$" #,##0.00'
            try:
                val = str(cell.value) if cell.value is not None else ""
                if len(val) > max_length: max_length = len(val)
            except: pass
            
        worksheet.column_dimensions[column_letter].width = min(max_length + 3, 50)
        
    worksheet.auto_filter.ref = worksheet.dimensions
    
    if nome_aba == 'Separação por Empenho':
        fill_cor1 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        fill_cor2 = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        empenho_atual = None
        usar_cor_cinza = False
        for row in range(2, worksheet.max_row + 1):
            celula_empenho = worksheet[f"F{row}"].value
            if celula_empenho != empenho_atual:
                usar_cor_cinza = not usar_cor_cinza
                empenho_atual = celula_empenho
            preenchimento_atual = fill_cor2 if usar_cor_cinza else fill_cor1
            for col_idx in range(1, worksheet.max_column + 1):
                celula = worksheet.cell(row=row, column=col_idx)
                celula.fill = preenchimento_atual
                celula.border = borda_fina

def selecionar_estoque():
    global caminho_estoque_global
    caminho = filedialog.askopenfilename(filetypes=[("PDF files", "*.pdf")])
    if caminho:
        caminho_estoque_global = caminho
        lbl_estoque.config(text=f"✅ {os.path.basename(caminho)}", fg="green")

def selecionar_pedidos():
    global caminho_pedidos_global
    caminho = filedialog.askopenfilename(filetypes=[("PDF files", "*.pdf")])
    if caminho:
        caminho_pedidos_global = caminho
        lbl_pedidos.config(text=f"✅ {os.path.basename(caminho)}", fg="green")

def selecionar_datas():
    global caminho_datas_global
    caminho = filedialog.askopenfilename(filetypes=[("Planilhas Excel/CSV", "*.xlsx;*.xls;*.csv")])
    if caminho:
        caminho_datas_global = caminho
        lbl_datas.config(text=f"✅ {os.path.basename(caminho)}", fg="green")

def executar():
    if not caminho_estoque_global or not caminho_pedidos_global:
        messagebox.showwarning("Aviso", "Selecione Estoque e Pedidos primeiro.")
        return
    
    salvar_na_nuvem = (modo_exportacao_var.get() == "nuvem")
    
    btn_executar.config(bg="orange", text="PROCESSANDO...", fg="black")
    root.update()
    
    try:
        progress_bar['value'] = 10
        atualizar_progresso(10, "CONECTANDO AO GOOGLE SHEETS...")
        planilha = obter_planilha_nuvem()
        
        progress_bar['value'] = 20
        atualizar_progresso(20, "LENDO HISTÓRICO DE ESTOQUE...")
        estoque_anterior = ler_estoque_anterior(planilha)
        
        progress_bar['value'] = 35
        atualizar_progresso(35, "EXTRAINDO ESTOQUE (PODE DEMORAR)...")
        estoque_atual = extrair_estoque_pdf(caminho_estoque_global)
        
        progress_bar['value'] = 50
        atualizar_progresso(50, "VERIFICANDO COMPRAS EM TRÂNSITO...")
        compras_transito = processar_auto_baixa_transito(planilha, estoque_atual, estoque_anterior, salvar_nuvem=salvar_na_nuvem)
        
        progress_bar['value'] = 60
        atualizar_progresso(60, "LENDO PLANILHA OPCIONAL...")
        infos = ler_planilha_datas(caminho_datas_global)
        
        progress_bar['value'] = 75
        atualizar_progresso(75, "EXTRAINDO PEDIDOS E INSUMOS...")
        demanda = extrair_pedidos_pdf(caminho_pedidos_global, estoque_atual, compras_transito)
        
        if not demanda: raise Exception("Nenhum empenho encontrado nos PDFs.")
        
        progress_bar['value'] = 85
        atualizar_progresso(85, "CRUZANDO E INJETANDO FÓRMULAS...")
        df_sep, df_comp, df_est = processar_cruzamento_dados(estoque_atual, demanda, infos, modo_prioridade_var.get(), compras_transito, salvar_na_nuvem)
        
        progress_bar['value'] = 95
        
        if salvar_na_nuvem:
            atualizar_progresso(95, "PINTANDO E ENVIANDO PARA A NUVEM...")
            exportar_google_sheets(planilha, df_sep, df_comp, df_est)
            progress_bar['value'] = 100
            messagebox.showinfo("Sucesso", "Dados e Formatação enviados para o Google Sheets com Sucesso!")
        else:
            atualizar_progresso(95, "SALVANDO EXCEL LOCALMENTE...")
            diretorio = os.path.dirname(caminho_pedidos_global)
            nome_saida = os.path.join(diretorio, "Relatorio_Teste_Local.xlsx")
            
            with pd.ExcelWriter(nome_saida, engine='openpyxl') as writer:
                df_sep.to_excel(writer, sheet_name='Separação por Empenho', index=False)
                df_comp.to_excel(writer, sheet_name='Resumo para Compras', index=False)
                df_est.to_excel(writer, sheet_name='Estoque Inicial', index=False)
                
                formatar_excel(writer, 'Separação por Empenho')
                formatar_excel(writer, 'Resumo para Compras')
                formatar_excel(writer, 'Estoque Inicial')
                
            progress_bar['value'] = 100
            messagebox.showinfo("Modo Teste - Sucesso", f"Planilha gerada salva na pasta:\n{nome_saida}\n\nA NUVEM NÃO FOI ALTERADA.")

    except Exception as e:
        messagebox.showerror("Erro", str(e))
    finally:
        btn_executar.config(text="RODAR INTELIGÊNCIA", bg="#0052cc", fg="white")
        progress_bar['value'] = 0

def atualizar_progresso(valor, texto):
    progress_bar['value'] = valor
    btn_executar.config(text=texto)
    root.update()

root = tk.Tk()
root.title("Gestor Inteligente de Empenhos Licitax")
root.geometry("550x620")
root.eval('tk::PlaceWindow . center')

tk.Label(root, text="1. PDF de Balanço de Estoque:", font=("Arial", 10, "bold")).pack(pady=(15, 2))
tk.Button(root, text="Selecionar Estoque (PDF)", command=selecionar_estoque).pack()
lbl_estoque = tk.Label(root, text="Nenhum PDF selecionado", fg="gray")
lbl_estoque.pack()

tk.Label(root, text="2. PDF de Pedidos de Venda:", font=("Arial", 10, "bold")).pack(pady=(15, 2))
tk.Button(root, text="Selecionar Pedidos (PDF)", command=selecionar_pedidos).pack()
lbl_pedidos = tk.Label(root, text="Nenhum PDF selecionado", fg="gray")
lbl_pedidos.pack()

tk.Label(root, text="3. Ficheiro VENDAS REALIZADAS (Opcional):", font=("Arial", 10, "bold")).pack(pady=(15, 2))
tk.Button(root, text="Selecionar Planilha", command=selecionar_datas).pack()
lbl_datas = tk.Label(root, text="Nenhuma Planilha selecionada", fg="gray")
lbl_datas.pack()

tk.Label(root, text="4. Como resolver empates?", font=("Arial", 10, "bold")).pack(pady=(10, 2))
modo_prioridade_var = tk.StringVar(value="unidades")
frame_opcoes = tk.Frame(root)
frame_opcoes.pack()
tk.Radiobutton(frame_opcoes, text="Falta de Unidades", variable=modo_prioridade_var, value="unidades").pack(anchor="w")
tk.Radiobutton(frame_opcoes, text="Falta de SKUs", variable=modo_prioridade_var, value="tipos").pack(anchor="w")

tk.Label(root, text="5. Destino do Relatório:", font=("Arial", 10, "bold", "underline")).pack(pady=(15, 2))
modo_exportacao_var = tk.StringVar(value="excel")
frame_exportacao = tk.Frame(root)
frame_exportacao.pack()
tk.Radiobutton(frame_exportacao, text="🗂️ Criar Excel Local (Modo de Teste)", variable=modo_exportacao_var, value="excel", fg="blue").pack(anchor="w")
tk.Radiobutton(frame_exportacao, text="☁️ Atualizar Google Sheets (Modo Oficial)", variable=modo_exportacao_var, value="nuvem", fg="red").pack(anchor="w")

progress_bar = ttk.Progressbar(root, maximum=100)
progress_bar.pack(pady=15, fill=tk.X, padx=40)

btn_executar = tk.Button(root, text="RODAR INTELIGÊNCIA", command=executar, bg="#0052cc", fg="white", font=("Arial", 12, "bold"))
btn_executar.pack(pady=5)

root.mainloop()